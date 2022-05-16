import os
import re
import shutil
import subprocess
import time
import openpyxl

cases_dir = "D:/workspace/firrtl_test_ise/synthesis_script/cases_dir.txt"
synthesis_output_dir = "D:/workspace/firrtl_test_ise/synthesis_script/output"
xst = "C:/Xilinx/14.7/ISE_DS/ISE/bin/nt64/xst.exe"
data_file = f'{synthesis_output_dir}/compare_data.xlsx'

def create_xst_file(case_name, xst_file_path, top_module):
    xst_content = f'set -tmpdir "xst/projnav.tmp"\n'
    xst_content += f'set -xsthdpdir "xst"\n'
    xst_content += 'run\n'
    xst_content += f'-ifn {case_name}.prj\n'
    xst_content += f'-ofn {case_name}\n'
    xst_content += ('-ofmt NGC\n'
                    '-p xc7vx690t-3-ffg1761\n')
    xst_content += f'-top {top_module}\n'
    xst_content += ('-opt_mode Speed\n'
                    '-opt_level 1\n'
                    '-power NO\n'
                    '-iuc NO\n'
                    '-keep_hierarchy No\n'
                    '-netlist_hierarchy As_Optimized\n'
                    '-rtlview Yes\n'
                    '-glob_opt AllClockNets\n'
                    '-read_cores YES\n'
                    '-write_timing_constraints NO\n'
                    '-cross_clock_analysis NO\n'
                    '-hierarchy_separator /\n'
                    '-bus_delimiter <>\n'
                    '-case Maintain\n'
                    '-slice_utilization_ratio 100\n'
                    '-bram_utilization_ratio 100\n'
                    '-dsp_utilization_ratio 100\n'
                    '-lc Auto\n'
                    '-reduce_control_sets Auto\n'
                    '-fsm_extract YES -fsm_encoding Auto\n'
                    '-safe_implementation No\n'
                    '-fsm_style LUT\n'
                    '-ram_extract Yes\n'
                    '-ram_style Auto\n'
                    '-rom_extract Yes\n'
                    '-shreg_extract YES\n'
                    '-rom_style Auto\n'
                    '-auto_bram_packing NO\n'
                    '-resource_sharing YES\n'
                    '-async_to_sync NO\n'
                    '-shreg_min_size 2\n'
                    '-use_dsp48 Auto\n'
                    '-iobuf YES\n'
                    '-max_fanout 100000\n'
                    '-bufg 32\n'
                    '-register_duplication YES\n'
                    '-register_balancing No\n'
                    '-optimize_primitives NO\n'
                    '-use_clock_enable Auto\n'
                    '-use_sync_set Auto\n'
                    '-use_sync_reset Auto\n'
                    '-iob Auto\n'
                    '-equivalent_register_removal YES\n'
                    '-slice_utilization_ratio_maxmargin 5\n')
    with open(xst_file_path, 'w') as xst_file:
        xst_file.write(xst_content)


def create_prj_file(case_name, prj_file_path):
    prj_content = f'verilog work "{case_name}.v"\n'
    with open(prj_file_path, 'w') as prj_file:
        prj_file.write(prj_content)


def run_xst(verilog_dir, case_name, new_case_name, top_module_name, subdir):
    case_path = f'{verilog_dir}/{subdir}/{case_name}.v'
    out_path = f'{synthesis_output_dir}/{new_case_name}/{subdir}'
    verilog_path = f'{out_path}/{new_case_name}.v'
    os.makedirs(out_path) if not os.path.exists(out_path) else None
    shutil.copy(case_path, verilog_path)
    xst_projnav_tmp_path = f'{out_path}/xst/projnav.tmp/'
    os.makedirs(xst_projnav_tmp_path) if not os.path.exists(xst_projnav_tmp_path) else None
    prj_file = f'{out_path}/{new_case_name}.prj'
    create_prj_file(new_case_name, prj_file)
    xst_file = f'{out_path}/{new_case_name}.xst'
    create_xst_file(new_case_name, xst_file, top_module_name)
    log_file = f'{out_path}/{new_case_name}.syr'
    os.chdir(out_path)
    xst_command = f'{xst} -intstyle ise -ifn "{xst_file}" -ofn "{log_file}"'
    xst_process = subprocess.Popen(xst_command, shell=True)
    start_time = time.time()
    while True:
        if xst_process.poll() is not None:
            break
        end_time = time.time()
        if round(end_time - start_time, 0) > 6000:
            xst_process.kill()
            print(f'Run case {new_case_name} out of time')
            break

def collect_timing_data(log_path, subdir, data_workbook, row_index):
    timing_data = data_workbook["timing data"]
    area_data = data_workbook["area data"]
    with open(log_path, 'r') as log_file:
        syn_log = log_file.read()
    #sequential_delay
    sequential_timing = re.findall('   Minimum period: (.*?)ns \(Maximum Frequency: (.*?)\)', syn_log)
    min_period = sequential_timing[0][0] if sequential_timing else -1
    #combinational delay
    combinational_timing = re.findall('   Maximum combinational path delay: (.*?)ns\n', syn_log)
    max_path_delay = combinational_timing[0] if combinational_timing else -1
    #slice register count
    slice_reg = re.findall(' Number of Slice Registers:  (.*)?  out of  866400  (.*)?%', syn_log)
    slice_reg_count = slice_reg[0][0] if slice_reg else -1
    #slice lut count
    slice_lut = re.findall(' Number of Slice LUTs:      (.*)?  out of  433200  (.*)?%', syn_log)
    slice_lut_count = slice_lut[0][0] if slice_lut else -1
    #LUT flip flop pairs
    lut_ffpair = re.findall(' Number of LUT Flip Flop pairs used:  (.*)?', syn_log)
    lut_ffpair_count = lut_ffpair[0] if lut_ffpair else -1
    #IO utilization
    io = re.findall(' Number of IOs:                    (.*)?', syn_log)
    io_count = io[0] if io else -1
    #BUFG/BUFCTRL
    bufg = re.findall(' Number of BUFG/BUFGCTRLs:           (.*)?  out of     32  (.*)?%', syn_log)
    bufg_count = bufg[0][0] if bufg else -1
    #DSP48E1s
    dsp = re.findall(' Number of DSP48E1s:                (.*)?  out of   3600   (.*)?%', syn_log)
    dsp_count = dsp[0][0] if dsp else -1
    #Block RAM/FIFO
    ram = re.findall(' Number of Block RAM/FIFO:           (.*)?  out of   1470   (.*)?%', syn_log)
    ram_count = ram[0][0] if ram else -1
    if subdir == "reference":
        ref_timing_list.append([min_period, max_path_delay])
        timing_data.cell(row=row_index, column=2, value=float(min_period))
        timing_data.cell(row=row_index, column=5, value=float(max_path_delay))
        area_data.cell(row=row_index, column=2, value=int(slice_reg_count))
        area_data.cell(row=row_index, column=5, value=int(slice_lut_count))
        area_data.cell(row=row_index, column=8, value=int(lut_ffpair_count))
        area_data.cell(row=row_index, column=11, value=int(dsp_count))
        area_data.cell(row=row_index, column=14, value=int(ram_count))
        area_data.cell(row=row_index, column=17, value=int(io_count))
        area_data.cell(row=row_index, column=20, value=int(bufg_count))
    elif subdir == "sbt":
        sbt_timing_list.append([min_period, max_path_delay])
        timing_data.cell(row=row_index, column=3, value=float(min_period))
        timing_data.cell(row=row_index, column=6, value=float(max_path_delay))
        area_data.cell(row=row_index, column=3, value=int(slice_reg_count))
        area_data.cell(row=row_index, column=6, value=int(slice_lut_count))
        area_data.cell(row=row_index, column=9, value=int(lut_ffpair_count))
        area_data.cell(row=row_index, column=12, value=int(dsp_count))
        area_data.cell(row=row_index, column=15, value=int(ram_count))
        area_data.cell(row=row_index, column=18, value=int(io_count))
        area_data.cell(row=row_index, column=21, value=int(bufg_count))
    data_workbook.save(data_file)



verilog_dir_list = []
try:
    with open(cases_dir, 'r') as verilog_dirs_file:
        verilog_dir_list = verilog_dirs_file.readlines()
except:
    print('Failed to open verilog source path')

if not os.path.exists(synthesis_output_dir):
    os.mkdir(synthesis_output_dir)
else:
    pass

ref_timing_list = []
sbt_timing_list = []
data_workbook = openpyxl.Workbook()
timing_data = data_workbook.create_sheet("timing data")
area_data = data_workbook.create_sheet("area data")
#write timing data table header
timing_data.cell(row=1, column=1, value='case name')
timing_data.cell(row=1, column=2, value='reference min period')
timing_data.cell(row=1, column=3, value='sbt min period')
timing_data.cell(row=1, column=5, value='reference max delay')
timing_data.cell(row=1, column=6, value='sbt max delay')
#write area data table header
area_data.cell(row=1, column=1, value='case name')
area_data.cell(row=1, column=2, value='reference Slice Register')
area_data.cell(row=1, column=3, value='sbt Slice Register')
area_data.cell(row=1, column=5, value='reference Slice LUT')
area_data.cell(row=1, column=6, value='sbt Slice LUT')
area_data.cell(row=1, column=8, value='reference LUT FF Pair')
area_data.cell(row=1, column=9, value='sbt LUT FF Pair')
area_data.cell(row=1, column=11, value='reference DSP')
area_data.cell(row=1, column=12, value='sbt DSP')
area_data.cell(row=1, column=14, value='reference BRAM/FIFO')
area_data.cell(row=1, column=15, value='sbt BRAM/FIFO')
area_data.cell(row=1, column=17, value='reference IO')
area_data.cell(row=1, column=18, value='sbt IO')
area_data.cell(row=1, column=20, value='reference BUFG/BUFGCTRL')
area_data.cell(row=1, column=21, value='sbt BUFG/BUFGCTRL')

row_index = 1
for verilog_info in verilog_dir_list:
    #verilog_dir = verilog_dir_list[0]
    row_index = row_index + 1
    timing_data = data_workbook["timing data"]
    area_data = data_workbook["area data"]
    verilog_info = verilog_info.strip().split()
    if len(verilog_info) != 2:
        print(f'Failed to read verilog direction and top module name!')
    verilog_dir = verilog_info[0]
    top_module = verilog_info[1]
    if not verilog_dir:
        print(f'{verilog_dir} is illegal!')
        continue
    if os.path.exists(verilog_dir):
        case_name = os.path.basename(verilog_dir)
        timing_data.cell(row=row_index, column=1, value=case_name)
        area_data.cell(row=row_index, column=1, value=case_name)
        new_case_name = case_name.replace('.', '_')
        top_module_name = top_module
        print(f'Synthesising {case_name}...')
        run_xst(verilog_dir, case_name, new_case_name, top_module_name, "reference")
        ref_log_path = f'{synthesis_output_dir}/{new_case_name}/reference/{new_case_name}.syr'
        collect_timing_data(ref_log_path, "reference", data_workbook, row_index)

        run_xst(verilog_dir, case_name, new_case_name, top_module_name, "sbt")
        sbt_log_path = f'{synthesis_output_dir}/{new_case_name}/sbt/{new_case_name}.syr'
        collect_timing_data(sbt_log_path, "sbt", data_workbook, row_index)


    else:
        print(f'{verilog_dir} does not exists!')
